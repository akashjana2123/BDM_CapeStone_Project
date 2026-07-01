"""
================================================================================
 BDM Capstone : Analytical Study of Ankita Cashew Processing (with Adrita)
 Mid-term analysis script  |  Akash Jana  |  Roll 23f2000990
--------------------------------------------------------------------------------
 This single file reproduces every number and every chart used in the mid-term
 report and the Excel workbook. Run it from the folder that contains the two
 source files:
        Sales Purchase 22-26 (1).xlsx      (sales & purchase ledgers)
        Stock Summary.xlsx                 (year-end stock records)

 It produces:
        results.json            -> every computed figure
        figures/figNN_*.png     -> every chart (Figure 1..N in the report)
        cleaning_log.txt        -> data-cleaning counts (duplicates, blanks, etc.)

 Requirements: pandas, numpy, statsmodels, matplotlib, openpyxl
        pip install pandas numpy statsmodels matplotlib openpyxl
================================================================================
"""
import os, re, json, datetime, warnings
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from statsmodels.tsa.seasonal import seasonal_decompose
from statsmodels.tsa.statespace.sarimax import SARIMAX
warnings.filterwarnings("ignore")

SALES_FILE = "Sales Purchase 22-26 (1).xlsx"
STOCK_FILE = "Stock Summary.xlsx"
FIGDIR = "figures"
os.makedirs(FIGDIR, exist_ok=True)
log_lines = []
def log(msg):
    print(msg); log_lines.append(str(msg))

# colours and number formatters used in all charts
C1, C2, C3, C4 = "#1f4e79", "#c55a11", "#548235", "#7030a0"
LAKH = FuncFormatter(lambda x, _: f"{x/1e5:.0f}")
CRORE = FuncFormatter(lambda x, _: f"{x/1e7:.1f}")
plt.rcParams.update({"font.size": 11, "figure.dpi": 130, "axes.grid": True,
                     "grid.alpha": .3, "axes.axisbelow": True})

# ==============================================================================
# STEP 1 : CLEAN THE SALES & PURCHASE LEDGERS
# ==============================================================================
# Each ledger sheet stacks several financial-year blocks. A real transaction row
# starts with a date in column 1; everything else (titles, addresses, the
# 'Date/Particulars' header, sub-totals and 'Closing Balance' lines) is noise.
SHEET_MAP = {
    "Adrita RCN Sale":      ("Adrita", "Sale",     "RCN"),
    "Adrita FCN Sale":      ("Adrita", "Sale",     "FCN"),
    "Adrita Purchase RCN":  ("Adrita", "Purchase", "RCN"),
    "Adrita Purchase FCN":  ("Adrita", "Purchase", "FCN"),
    "Ankita RCN Sale":      ("Ankita", "Sale",     "RCN"),
    "Ankita FCN Sale":      ("Ankita", "Sale",     "FCN"),
    "Ankita Purchase RCN":  ("Ankita", "Purchase", "RCN"),
    "Ankita Purchase FCN":  ("Ankita", "Purchase", "FCN"),
    "Ankita Export":        ("Ankita", "Export",   "FCN"),
}

def clean_ledgers(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows, scanned, closing_removed, blank_skipped = [], 0, 0, 0
    for sheet, (comp, flow, goods) in SHEET_MAP.items():
        ws = wb[sheet]
        for r in range(1, ws.max_row + 1):
            scanned += 1
            d = ws.cell(r, 1).value
            if not isinstance(d, (datetime.datetime, datetime.date)):
                blank_skipped += 1                    # title / header / blank row
                continue
            name = ws.cell(r, 3).value
            if name and "Closing Balance" in str(name):
                closing_removed += 1                  # ledger closing line
                continue
            debit, credit = ws.cell(r, 6).value, ws.cell(r, 7).value
            amount = credit if flow in ("Sale", "Export") else debit
            if amount is None:                        # use whichever column is filled
                amount = credit if credit is not None else debit
            if amount is None:
                blank_skipped += 1
                continue
            rows.append(dict(Company=comp, Flow=flow, Goods=goods,
                             Date=pd.Timestamp(d),
                             Counterparty=str(name).strip() if name else "",
                             Amount=float(amount)))
    df = pd.DataFrame(rows)
    before = len(df)
    df = df.drop_duplicates()                          # remove exact duplicate rows
    dupes = before - len(df)
    df["FY"] = df.Date.apply(lambda x: x.year if x.month >= 4 else x.year - 1)
    df["FYlabel"] = df.FY.apply(lambda y: f"FY{str(y)[2:]}-{str(y+1)[2:]}")
    df["Month"] = df.Date.dt.strftime("%b")
    df["YearMonth"] = df.Date.dt.to_period("M").astype(str)
    log(f"[CLEANING] cells scanned in ledgers ............ {scanned}")
    log(f"[CLEANING] non-transaction rows skipped ........ {blank_skipped}")
    log(f"[CLEANING] 'Closing Balance' lines removed ..... {closing_removed}")
    log(f"[CLEANING] exact duplicate rows removed ........ {dupes}")
    log(f"[CLEANING] clean transactions kept ............. {len(df)}")
    log(f"[CLEANING] missing Amount values after clean ... {int(df.Amount.isna().sum())}")
    return df, dict(scanned=scanned, blank_skipped=blank_skipped,
                    closing_removed=closing_removed, duplicates=dupes,
                    kept=len(df), missing=int(df.Amount.isna().sum()))

# ==============================================================================
# STEP 2 : CLEAN THE STOCK SUMMARY (monthly inward/outward, in KG)
# ==============================================================================
MONTHS = {"April":4,"May":5,"June":6,"July":7,"August":8,"September":9,
          "October":10,"November":11,"December":12,"January":1,"February":2,"March":3}

def kg(s):
    if not isinstance(s, str):
        return None
    m = re.match(r"\s*(-?[\d,]+(?:\.\d+)?)\s*KG", s, re.I)
    return float(m.group(1).replace(",", "")) if m else None

def grade(item):
    u = item.upper()
    if "RAW CASHEW" in u or "NUT IN SHELL" in u: return "RCN"
    if "WHOLE" in u: return "Wholes"
    if "BROKEN" in u: return "Broken"
    return None

def clean_stock(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    monthly, closing = [], []
    for sheet in wb.sheetnames:
        comp = "Ankita" if "Ankita" in sheet else "Adrita"
        fy = int("20" + sheet.split()[-1])
        ws = wb[sheet]
        # (a) closing-balance block -> used for ABC classification
        start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 2).value == "Quantity" and ws.cell(r, 4).value == "Value":
                start = r + 1; break
        if start:
            for r in range(start, ws.max_row + 1):
                nm = ws.cell(r, 1).value
                if not isinstance(nm, str): continue
                if nm.strip() == "Grand Total": break
                v = ws.cell(r, 4).value
                if isinstance(v, (int, float)):
                    closing.append(dict(Company=comp, Year=f"20{sheet.split()[-1]}",
                                        Item=nm.strip(), Value=float(v)))
        # (b) per-item monthly inward / outward tables
        cur = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            nxt = ws.cell(r + 1, 1).value if r < ws.max_row else None
            if isinstance(a, str) and nxt == "Monthly Summary":
                cur = a.strip()
            if isinstance(a, str) and a.strip() in MONTHS and cur:
                g = grade(cur)
                if g:
                    inq = kg(ws.cell(r, 2).value) or 0
                    inval = ws.cell(r, 3).value or 0
                    outq = kg(ws.cell(r, 4).value) or 0
                    m = MONTHS[a.strip()]; yr = fy if m >= 4 else fy + 1
                    monthly.append(dict(Company=comp,
                                        FYlabel=f"FY{str(fy)[2:]}-{str(fy+1)[2:]}",
                                        Date=pd.Timestamp(year=yr, month=m, day=1),
                                        Cat=g, In_KG=inq,
                                        In_Val=float(inval) if isinstance(inval,(int,float)) else 0,
                                        Out_KG=outq))
    mv = pd.DataFrame(monthly)
    cb = pd.DataFrame(closing)
    log(f"[CLEANING] stock monthly rows parsed ........... {len(mv)}")
    log(f"[CLEANING] stock closing-balance items ......... {len(cb)}")
    return mv, cb

# ==============================================================================
# STEP 3 : DESCRIPTIVE STATISTICS (all variables used in the analysis)
# ==============================================================================
def describe(series):
    s = series.dropna()
    return dict(count=int(s.count()), mean=float(s.mean()), median=float(s.median()),
                std=float(s.std()), min=float(s.min()), max=float(s.max()),
                total=float(s.sum()))

# ==============================================================================
# STEP 4 : DEMAND FORECASTING  (decomposition + linear + ARIMA + RMSE/MAE)
# ==============================================================================
def rmse(a, b): return float(np.sqrt(np.mean((np.array(a) - np.array(b)) ** 2)))
def mae(a, b):  return float(np.mean(np.abs(np.array(a) - np.array(b))))

def forecasting(df):
    sale = df[df.Flow.isin(["Sale", "Export"])]
    ts = sale.groupby(sale.Date.dt.to_period("M")).Amount.sum()
    ts.index = ts.index.to_timestamp()
    ts = ts.asfreq("MS").fillna(0)
    ts = ts[ts.index <= "2026-01-01"]               # drop the partial last month
    out = {"series": {d.strftime("%Y-%m"): float(v) for d, v in ts.items()}}

    # straight-line (linear) trend baseline:  y = a + b*t
    x = np.arange(len(ts))
    b, a = np.polyfit(x, ts.values, 1)
    out["lin_slope"], out["lin_intercept"] = float(b), float(a)
    out["lin_trend"] = {d.strftime("%Y-%m"): float(a + b * i) for i, d in enumerate(ts.index)}

    # decomposition into trend + season
    dec = seasonal_decompose(ts, model="additive", period=12)
    out["seasonal"] = {d.strftime("%b"): float(v) for d, v in zip(ts.index[:12], dec.seasonal.values[:12])}
    out["trend_comp"] = {d.strftime("%Y-%m"): (None if np.isnan(t) else float(t))
                         for d, t in zip(ts.index, dec.trend.values)}

    # VALIDATION: hold out the last 6 months, forecast them, compare
    n, h = len(ts), 6
    train, test = ts.iloc[:n-h], ts.iloc[n-h:]
    sar = SARIMAX(train, order=(1,1,1), seasonal_order=(0,1,1,12),
                  enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
    sar_pred = sar.forecast(h)
    out["sarima_rmse"], out["sarima_mae"] = rmse(test, sar_pred), mae(test, sar_pred)
    lin_pred = a + b * np.arange(n-h, n)
    out["lin_rmse"], out["lin_mae"] = rmse(test, lin_pred), mae(test, lin_pred)
    naive = ts.shift(12).iloc[n-h:]
    out["naive_rmse"], out["naive_mae"] = rmse(test, naive), mae(test, naive)
    out["test_actual"] = {d.strftime("%Y-%m"): float(v) for d, v in test.items()}
    out["sarima_test_pred"] = {d.strftime("%Y-%m"): float(v) for d, v in sar_pred.items()}
    out["lin_test_pred"] = {d.strftime("%Y-%m"): float(v) for d, v in zip(test.index, lin_pred)}

    # FUTURE: refit on all data, forecast the next 4 months
    full = SARIMAX(ts, order=(1,1,1), seasonal_order=(0,1,1,12),
                   enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
    f = full.get_forecast(4); fm, ci = f.predicted_mean, f.conf_int(alpha=0.2)
    out["forecast"] = {d.strftime("%Y-%m"): {"mean": float(v),
                       "lo": float(max(0, ci.iloc[i,0])), "hi": float(ci.iloc[i,1])}
                       for i, (d, v) in enumerate(fm.items())}
    log(f"[FORECAST] ARIMA  RMSE={out['sarima_rmse']:.0f}  MAE={out['sarima_mae']:.0f}")
    log(f"[FORECAST] Linear RMSE={out['lin_rmse']:.0f}  MAE={out['lin_mae']:.0f}")
    log(f"[FORECAST] Naive  RMSE={out['naive_rmse']:.0f}  MAE={out['naive_mae']:.0f}")
    return out, ts

# ==============================================================================
# STEP 5 : INVENTORY YIELD  (yield % per cycle, variance vs target, Z-score) + ABC
# ==============================================================================
TARGET_YIELD = 22.0       # normal cashew outturn used as the target

def yield_analysis(mv, cb):
    # per company-year: RCN bought (inward) and Whole+Broken kernel sold (outward)
    agg = mv.groupby(["Company", "FYlabel", "Cat"]).agg(
        In_KG=("In_KG", "sum"), Out_KG=("Out_KG", "sum")).reset_index()
    rows = []
    for (c, fy), g in agg.groupby(["Company", "FYlabel"]):
        d = {r.Cat: r for r in g.itertuples()}
        rcn = d["RCN"].In_KG if "RCN" in d else 0
        wo = d["Wholes"].Out_KG if "Wholes" in d else 0
        br = d["Broken"].Out_KG if "Broken" in d else 0
        kern = wo + br
        rows.append(dict(Company=c, Year=fy, RCN_Purchased=rcn, Wholes_Sold=wo,
                         Broken_Sold=br, Kernel_Out=kern,
                         Yield_pct=round(100*kern/rcn, 2) if rcn else None,
                         Whole_share=round(100*wo/kern, 1) if kern else None))
    Y = pd.DataFrame(rows).dropna(subset=["Yield_pct"]).reset_index(drop=True)
    mu, sd = Y.Yield_pct.mean(), Y.Yield_pct.std()
    Y["Target"] = TARGET_YIELD
    Y["Variance"] = (Y.Yield_pct - TARGET_YIELD).round(2)
    Y["Zscore"] = ((Y.Yield_pct - mu) / sd).round(2)
    Y["Outlier"] = np.where(Y.Zscore.abs() > 1.5, "Yes", "No")

    comp = Y.groupby("Company").agg(RCN=("RCN_Purchased","sum"), Kern=("Kernel_Out","sum"),
                                    W=("Wholes_Sold","sum"), B=("Broken_Sold","sum")).reset_index()
    comp["Recovery_pct"] = (comp.Kern / comp.RCN * 100).round(2)
    comp["Whole_share"] = (comp.W / (comp.W + comp.B) * 100).round(1)

    # ABC classification of stock items by value
    def group(it):
        u = it.upper()
        if "RAW CASHEW" in u or "NUT IN SHELL" in u: return "Raw Cashew Nut (RCN)"
        if "WHOLE" in u: return "Whole Kernel"
        if "BROKEN" in u: return "Broken Kernel"
        if "REJECT" in u or "MIX" in u: return "Rejection / Mix"
        if any(k in u for k in ["TIN CAN","CARTOON","PACK","POUCH","PAUCH"]): return "Packing Material"
        if any(k in u for k in ["MACHIN","CAPITAL","COMPUTER","CAMERA","BATTRY","BRAKET","ELECTRIC"]): return "Plant & Equipment"
        if "BUILDING" in u: return "Building Material"
        return "Other"
    cb = cb.copy(); cb["AbsValue"] = cb.Value.abs(); cb["Group"] = cb.Item.apply(group)
    g = cb.groupby("Group").AbsValue.sum().sort_values(ascending=False)
    tot = g.sum(); cum = g.cumsum() / tot * 100
    abc = []
    for name, val in g.items():
        cp = float(cum[name]); cls = "A" if cp <= 70 else ("B" if cp <= 90 else "C")
        abc.append(dict(Group=name, Value=float(val), Pct=round(val/tot*100,1),
                        CumPct=round(cp,1), Class=cls))
    log(f"[YIELD] mean yield={mu:.1f}%  sd={sd:.1f}%  outliers={list(Y[Y.Outlier=='Yes'].Company+' '+Y[Y.Outlier=='Yes'].Year)}")
    return Y, comp, abc, float(mu), float(sd), float(tot)

# ==============================================================================
# STEP 6 : FIGURES  (grouped problem-wise, named Figure 1..N)
# ==============================================================================
def D(keys): return [pd.Timestamp(k + "-01") for k in keys]

def make_figures(df, fc, ts, Y, comp, abc, mv):
    sale = df[df.Flow.isin(["Sale", "Export"])]
    # ---- OVERVIEW ----
    asp = df.assign(Side=np.where(df.Flow=="Purchase","Purchase","Sales")).groupby(
        ["FYlabel","Side"]).Amount.sum().unstack().fillna(0)
    fys = list(asp.index)
    fig, ax = plt.subplots(figsize=(7.5,4.2)); x=np.arange(len(fys)); w=.38
    ax.bar(x-w/2, asp.Sales, w, label="Sales", color=C1)
    ax.bar(x+w/2, asp.Purchase, w, label="Purchase", color=C2)
    ax.set_xticks(x); ax.set_xticklabels(fys); ax.yaxis.set_major_formatter(CRORE)
    ax.set_ylabel("₹ Crore"); ax.set_title("Figure 1: Annual Sales vs Purchase")
    ax.legend(); plt.tight_layout(); plt.savefig(f"{FIGDIR}/fig01_annual.png"); plt.close()

    # ---- DEMAND & PRICE FORECASTING ----
    s = fc["series"]; di = D(list(s))
    fig, ax = plt.subplots(3,1,figsize=(9,7),sharex=True)
    ax[0].plot(di, list(s.values()), color=C1, lw=1.4); ax[0].set_ylabel("Sales"); ax[0].yaxis.set_major_formatter(LAKH)
    ax[0].set_title("Figure 2: Monthly Sales split into Trend and Season (₹ Lakh)")
    tr = fc["trend_comp"]; ax[1].plot(di, [np.nan if tr[k] in (None,"None") else tr[k] for k in s], color=C2, lw=2)
    ax[1].set_ylabel("Trend"); ax[1].yaxis.set_major_formatter(LAKH)
    seas = fc["seasonal"]; ax[2].plot(di, [seas.get(pd.Timestamp(k+'-01').strftime('%b'),0) for k in s], color=C3, lw=1.4)
    ax[2].set_ylabel("Season"); ax[2].yaxis.set_major_formatter(LAKH)
    plt.tight_layout(); plt.savefig(f"{FIGDIR}/fig02_decomposition.png"); plt.close()

    months = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
    sm = sale.groupby("Month").Amount.sum()
    fig, ax = plt.subplots(figsize=(8,4))
    vals=[sm.get(m,0) for m in months]; bars=ax.bar(months, vals, color=C3)
    bars[int(np.argmax(vals))].set_color(C2)
    ax.yaxis.set_major_formatter(CRORE); ax.set_ylabel("₹ Crore")
    ax.set_title("Figure 3: Total Sales by Month (all years pooled)")
    plt.tight_layout(); plt.savefig(f"{FIGDIR}/fig03_seasonality.png"); plt.close()

    # Figure 4 : VALIDATION (holdout test, actual vs models) -- shown BEFORE prediction
    ta, sp, lp = fc["test_actual"], fc["sarima_test_pred"], fc["lin_test_pred"]
    td = D(list(ta))
    fig, ax = plt.subplots(figsize=(8,4))
    ax.plot(td, list(ta.values()), color=C1, marker="o", lw=1.8, label="Actual")
    ax.plot(td, list(sp.values()), color=C2, marker="s", ls="--", lw=1.8, label="ARIMA prediction")
    ax.plot(td, list(lp.values()), color="grey", marker="^", ls=":", lw=1.6, label="Linear prediction")
    ax.yaxis.set_major_formatter(LAKH); ax.set_ylabel("Sales (₹ Lakh)")
    ax.set_title("Figure 4: Model Validation on 6-Month Holdout (actual vs predicted)")
    ax.legend(fontsize=8); plt.xticks(rotation=45, fontsize=8); plt.tight_layout()
    plt.savefig(f"{FIGDIR}/fig04_validation.png"); plt.close()

    # Figure 5 : accuracy bars
    fig, ax = plt.subplots(figsize=(7,4))
    m=["ARIMA","Linear","Naive"]; rm=[fc["sarima_rmse"],fc["lin_rmse"],fc["naive_rmse"]]
    ma=[fc["sarima_mae"],fc["lin_mae"],fc["naive_mae"]]; xx=np.arange(3); w=.38
    ax.bar(xx-w/2, rm, w, label="RMSE", color=C1); ax.bar(xx+w/2, ma, w, label="MAE", color=C2)
    ax.set_xticks(xx); ax.set_xticklabels(m); ax.yaxis.set_major_formatter(LAKH); ax.set_ylabel("Error (₹ Lakh)")
    ax.set_title("Figure 5: Forecast Accuracy (lower is better)"); ax.legend()
    plt.tight_layout(); plt.savefig(f"{FIGDIR}/fig05_accuracy.png"); plt.close()

    # Figure 6 : future forecast
    fig, ax = plt.subplots(figsize=(9.5,4))
    ax.plot(di, list(s.values()), color=C1, lw=1.4, marker="o", ms=3, label="Actual sales")
    ff=fc["forecast"]; fd=D(list(ff))
    ax.plot(fd, [v["mean"] for v in ff.values()], color=C4, lw=2, marker="D", label="ARIMA forecast")
    ax.fill_between(fd, [v["lo"] for v in ff.values()], [v["hi"] for v in ff.values()], color=C4, alpha=.15)
    ax.yaxis.set_major_formatter(LAKH); ax.set_ylabel("Sales (₹ Lakh)")
    ax.set_title("Figure 6: Sales Forecast for the Next 4 Months")
    ax.legend(fontsize=8); plt.xticks(rotation=45, fontsize=8); plt.tight_layout()
    plt.savefig(f"{FIGDIR}/fig06_forecast.png"); plt.close()

    # Figure 7 : RCN price trend
    rcn = mv[(mv.Cat=="RCN") & (mv.In_KG>0) & (mv.In_Val>0)].copy()
    pr = rcn.groupby("Date").apply(lambda g: g.In_Val.sum()/g.In_KG.sum()).sort_index()
    fig, ax = plt.subplots(figsize=(9,3.8))
    ax.plot(pr.index, pr.values, color=C2, lw=1.6, marker="o", ms=3)
    ax.set_ylabel("₹ per KG"); ax.set_title("Figure 7: Raw Cashew Nut Buying Price per KG")
    plt.xticks(rotation=45, fontsize=8); plt.tight_layout(); plt.savefig(f"{FIGDIR}/fig07_rcn_price.png"); plt.close()

    # ---- INVENTORY & YIELD ----
    # Figure 8 : stock movement
    rin = mv[mv.Cat=="RCN"].groupby("Date").In_KG.sum()
    kout = mv[mv.Cat.isin(["Wholes","Broken"])].groupby("Date").Out_KG.sum()
    idx = sorted(set(rin.index) | set(kout.index))
    fig, ax = plt.subplots(figsize=(9.5,4))
    ax.bar(idx, [rin.get(d,0) for d in idx], width=20, color=C1, label="RCN bought (in)", alpha=.8)
    ax.bar(idx, [-kout.get(d,0) for d in idx], width=20, color=C2, label="Kernel sold (out)", alpha=.8)
    ax.axhline(0, color="black", lw=.6); ax.set_ylabel("KG (in up / out down)")
    ax.set_title("Figure 8: Monthly Stock Movement — Raw Nut In vs Kernel Out")
    ax.legend(fontsize=8); plt.xticks(rotation=45, fontsize=8); plt.tight_layout()
    plt.savefig(f"{FIGDIR}/fig08_movement.png"); plt.close()

    # Figure 9 : yield vs target
    lbl = (Y.Company.str[:2] + "-" + Y.Year.str[2:]).tolist()
    fig, ax = plt.subplots(figsize=(8.5,4.2))
    cols = [C3 if v >= 0 else C2 for v in Y.Variance]
    ax.bar(lbl, Y.Yield_pct, color=cols)
    ax.axhline(TARGET_YIELD, ls="--", color="black", lw=1.2, label=f"Target {TARGET_YIELD:.0f}%")
    for i,(yv,o_) in enumerate(zip(Y.Yield_pct, Y.Outlier)):
        ax.text(i, yv+0.3, f"{yv:.1f}", ha="center", fontsize=8, fontweight="bold" if o_=="Yes" else "normal")
    ax.set_ylabel("Yield %"); ax.set_title("Figure 9: Batch Yield vs Target")
    ax.legend(); plt.xticks(rotation=45, fontsize=8); plt.tight_layout()
    plt.savefig(f"{FIGDIR}/fig09_yield.png"); plt.close()

    # Figure 10 : z-score
    fig, ax = plt.subplots(figsize=(8.5,4))
    cols = ["#c0392b" if o_=="Yes" else C1 for o_ in Y.Outlier]
    ax.bar(lbl, Y.Zscore, color=cols)
    ax.axhline(1.5, ls="--", color="grey"); ax.axhline(-1.5, ls="--", color="grey")
    ax.set_ylabel("Z-score"); ax.set_title("Figure 10: Yield Z-score (beyond ±1.5 flagged)")
    plt.xticks(rotation=45, fontsize=8); plt.tight_layout(); plt.savefig(f"{FIGDIR}/fig10_zscore.png"); plt.close()

    # Figure 11 : ABC Pareto
    A = pd.DataFrame(abc); cmap = {"A":C2,"B":C1,"C":"#999999"}
    fig, ax = plt.subplots(figsize=(9,4.4))
    ax.bar(A.Group, A.Value, color=[cmap[c] for c in A["Class"]])
    ax.yaxis.set_major_formatter(CRORE); ax.set_ylabel("Value (₹ Cr)")
    ax2 = ax.twinx(); ax2.plot(A.Group, A.CumPct, color="black", marker="o", lw=1.4)
    ax2.set_ylabel("Cumulative %"); ax2.set_ylim(0,105); ax2.grid(False)
    ax.set_title("Figure 11: ABC Classification of Stock (Pareto)")
    plt.xticks(rotation=25, ha="right", fontsize=8); plt.tight_layout()
    plt.savefig(f"{FIGDIR}/fig11_abc.png"); plt.close()

    # ---- REVENUE / BUYERS ----
    # Figure 12 : revenue mix
    gmix = sale.groupby("Goods").Amount.sum()
    fig, ax = plt.subplots(figsize=(5.5,5))
    ax.pie([gmix.get("FCN",0), gmix.get("RCN",0)], labels=["Finished Kernels (FCN)","Raw Cashew (RCN)"],
           autopct="%1.1f%%", colors=[C1,C2], startangle=90, wedgeprops=dict(width=.45, edgecolor="w"))
    ax.set_title("Figure 12: Revenue Mix — Kernels vs Raw Nut"); plt.tight_layout()
    plt.savefig(f"{FIGDIR}/fig12_mix.png"); plt.close()

    # Figure 13 : top buyers
    tb = sale.groupby("Counterparty").Amount.sum().sort_values(ascending=False).head(8)[::-1]
    fig, ax = plt.subplots(figsize=(8,4.2))
    ax.barh([n[:28] for n in tb.index], tb.values, color=C1)
    ax.xaxis.set_major_formatter(CRORE); ax.set_xlabel("Sales (₹ Crore)")
    ax.set_title("Figure 13: Top Buyers by Revenue"); plt.tight_layout()
    plt.savefig(f"{FIGDIR}/fig13_buyers.png"); plt.close()
    log(f"[FIGURES] 13 figures written to ./{FIGDIR}/")

# ==============================================================================
# MAIN
# ==============================================================================
def main():
    df, clean_counts = clean_ledgers(SALES_FILE)
    mv, cb = clean_stock(STOCK_FILE)

    sale = df[df.Flow.isin(["Sale", "Export"])]
    pur = df[df.Flow == "Purchase"]
    rcn = mv[(mv.Cat=="RCN") & (mv.In_KG>0) & (mv.In_Val>0)].copy()
    rcn["Rate"] = rcn.In_Val / rcn.In_KG

    # descriptive statistics for every variable used downstream
    stats = {
        "Sale amount (₹)":            describe(sale.Amount),
        "Purchase amount (₹)":        describe(pur.Amount),
        "Monthly sales (₹)":          describe(sale.groupby(sale.Date.dt.to_period("M")).Amount.sum()),
        "RCN buying price (₹/KG)":    describe(rcn.Rate),
        "Monthly RCN bought (KG)":    describe(mv[mv.Cat=="RCN"].groupby("Date").In_KG.sum()),
        "Monthly kernel sold (KG)":   describe(mv[mv.Cat.isin(["Wholes","Broken"])].groupby("Date").Out_KG.sum()),
    }

    fc, ts = forecasting(df)
    Y, comp, abc, ymu, ysd, abctot = yield_analysis(mv, cb)
    # yield % added to stats once Y exists
    stats["Yield per cycle (%)"] = describe(Y.Yield_pct)

    make_figures(df, fc, ts, Y, comp, abc, mv)

    results = dict(
        clean_counts=clean_counts,
        totals=dict(total_sales=float(sale.Amount.sum()), total_purchase=float(pur.Amount.sum()),
                    n_sales=int(len(sale)), n_purchase=int(len(pur)),
                    date_min=str(df.Date.min().date()), date_max=str(df.Date.max().date())),
        descriptive=stats,
        company_flow=df.groupby(["Company","Flow"]).Amount.agg(["count","sum"]).reset_index().to_dict("records"),
        sales_by_goods={k: float(v) for k, v in sale.groupby("Goods").Amount.sum().items()},
        forecast=fc,
        rcn_price_yearly={int(y): round(float(v),2) for y,v in rcn.groupby(rcn.Date.dt.year).apply(
            lambda g: g.In_Val.sum()/g.In_KG.sum()).items()},
        yield_batches=Y.to_dict("records"), yield_company=comp.to_dict("records"),
        yield_mean=round(ymu,2), yield_sd=round(ysd,2), yield_target=TARGET_YIELD,
        abc=abc, abc_total=abctot,
        top_buyers=[{"name": k, "amount": float(v)} for k, v in
                    sale.groupby("Counterparty").Amount.sum().sort_values(ascending=False).head(10).items()],
        buyer_concentration_top5=float(sale.groupby("Counterparty").Amount.sum().sort_values(
            ascending=False).head(5).sum() / sale.Amount.sum() * 100),
    )
    json.dump(results, open("results.json", "w"), indent=2, default=str)

    # annual.json : sales & purchase per financial year (used by the report builder)
    annual = []
    for fy in sorted(df.FYlabel.unique()):
        g = df[df.FYlabel == fy]
        annual.append(dict(
            fy=fy,
            sales=float(g[g.Flow.isin(["Sale", "Export"])].Amount.sum()),
            purchase=float(g[g.Flow == "Purchase"].Amount.sum()),
        ))
    json.dump(annual, open("annual.json", "w"), indent=2)

    open("cleaning_log.txt", "w").write("\n".join(log_lines))
    log("[DONE] results.json + annual.json + cleaning_log.txt + figures/ written.")


if __name__ == "__main__":
    main()
