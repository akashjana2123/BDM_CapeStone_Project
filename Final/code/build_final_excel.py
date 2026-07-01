import json, datetime
import pandas as pd, numpy as np, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule

RF=json.load(open('results_final.json')); RM=json.load(open('results.json')); ANN=json.load(open('annual.json'))
bd=pd.read_pickle('buyers.pkl'); A=pd.read_pickle('anomdf.pkl'); recon=pd.read_pickle('recon.pkl')
SM={'Adrita RCN Sale':('Adrita','Sale','RCN'),'Adrita FCN Sale':('Adrita','Sale','FCN'),'Adrita Purchase RCN':('Adrita','Purchase','RCN'),'Adrita Purchase FCN':('Adrita','Purchase','FCN'),'Ankita RCN Sale':('Ankita','Sale','RCN'),'Ankita FCN Sale':('Ankita','Sale','FCN'),'Ankita Purchase RCN':('Ankita','Purchase','RCN'),'Ankita Purchase FCN':('Ankita','Purchase','FCN'),'Ankita Export':('Ankita','Export','FCN')}
wb0=openpyxl.load_workbook('Sales Purchase 22-26 (1).xlsx',data_only=True);rows=[]
for sh,(c,f,g) in SM.items():
    ws=wb0[sh]
    for r in range(1,ws.max_row+1):
        d=ws.cell(r,1).value
        if not isinstance(d,(datetime.datetime,datetime.date)):continue
        nm=ws.cell(r,3).value
        if nm and 'Closing Balance' in str(nm):continue
        deb,cred=ws.cell(r,6).value,ws.cell(r,7).value
        amt=cred if f in('Sale','Export') else deb
        if amt is None: amt=cred if cred is not None else deb
        if amt is None: continue
        rows.append(dict(Company=c,Flow=f,Goods=g,Date=pd.Timestamp(d),Counterparty=str(nm).strip() if nm else '',Amount=float(amt)))
df=pd.DataFrame(rows).drop_duplicates();df['FY']=df.Date.apply(lambda x:x.year if x.month>=4 else x.year-1);df['FYlabel']=df.FY.apply(lambda y:f"FY{str(y)[2:]}-{str(y+1)[2:]}")

TEAL='0F766E';AMBER='B45309';ROSE='BE123C';GREEN='15803D';W='FFFFFF'
hf=Font(name='Arial',bold=True,color=W,size=11);hfill=PatternFill('solid',fgColor=TEAL)
title=Font(name='Arial',bold=True,size=14,color=TEAL);sub=Font(name='Arial',bold=True,size=11,color=AMBER)
bd_=Border(*[Side(style='thin',color='BBBBBB')]*4);blue=Font(color='0000FF');ital=Font(italic=True,size=9,color='888888');center=Alignment('center',vertical='center')
def sh(ws,row,c1,c2):
    for c in range(c1,c2+1):
        x=ws.cell(row,c);x.font=hf;x.fill=hfill;x.alignment=center;x.border=bd_
def dim(ws,d):
    for k,v in d.items(): ws.column_dimensions[k].width=v
def styleCol(s,hexcol,line=False):
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    s.graphicalProperties.solidFill=hexcol
wb=Workbook()

# README
ws=wb.active;ws.title='README';ws.sheet_view.showGridLines=False
ws['B2']='Analytical Study of Ankita Cashew Processing';ws['B2'].font=title
ws['B3']='BDM Capstone — FINAL Analysis Workbook (native Excel charts)';ws['B3'].font=sub
ws['B4']='Akash Jana | Roll 23f2000990 | IIT Madras BS Degree'
for i,n in enumerate(['','All charts here are native Excel charts (editable; right-click > Edit Data).','Reproduced by analysis_final.py.','',
 'Sheets: Cleaning Log, Transactions, Descriptive Stats, Annual Summary, Forecast, Yield,',
 'P&L Model (live formulas), Sensitivity (colour scale), Buyer Risk Model, Buyer Tiers,',
 'Anomalies. Blue = inputs/assumptions, black = formulas. ₹ amounts, KG weights.']):
    ws.cell(5+i,2,n)
dim(ws,{'A':2,'B':95})

# Cleaning Log
ws=wb.create_sheet('Cleaning Log');ws['A1']='Data Cleaning Log';ws['A1'].font=title
ws.append([]);ws.append(['Step','Count']);sh(ws,3,1,2);cc=RF['clean_counts']
for i,(s,c) in enumerate([('Cells scanned',cc['scanned']),('Non-data rows skipped',cc['skipped']),('Exact duplicates removed',cc['duplicates']),('Clean transactions kept',cc['kept'])]):
    ws.cell(4+i,1,s);x=ws.cell(4+i,2,c);x.font=blue;x.number_format='#,##0';ws.cell(4+i,1).border=bd_;x.border=bd_
ws.cell(9,1,'Duplicate examples: ₹70,000 cash sale (Ankita 31-Mar-2023); ₹25,00,000 RCN purchase (Ankita 30-Jul-2024).').font=ital
dim(ws,{'A':46,'B':12})

# Transactions
ws=wb.create_sheet('Transactions')
d1=df[['Company','Flow','Goods','Date','Counterparty','Amount','FYlabel']].sort_values(['Company','Date'])
ws.append(['Company','Flow','Goods','Date','Counterparty','Amount (₹)','Financial Year']);sh(ws,1,1,7)
for _,x in d1.iterrows(): ws.append([x.Company,x.Flow,x.Goods,x.Date.date(),x.Counterparty,float(x.Amount),x.FYlabel])
for rr in range(2,ws.max_row+1): ws.cell(rr,6).number_format='#,##0';ws.cell(rr,4).number_format='dd-mmm-yy'
ws.freeze_panes='A2';ws.auto_filter.ref=f"A1:G{ws.max_row}";dim(ws,{'A':9,'B':10,'C':7,'D':12,'E':42,'F':14,'G':13})

# Descriptive Stats
ws=wb.create_sheet('Descriptive Stats');ws['A1']='Descriptive Statistics — all variables';ws['A1'].font=title
ws.append([]);ws.append(['Variable','Count','Mean','Median','SD','Min','Max']);sh(ws,3,1,7)
order=["Sale amount (₹)","Purchase amount (₹)","RCN price (₹/KG)","Buyer revenue (₹)","Buyer orders (count)","Buyer risk score (0-100)","Contribution per kg kernel (₹)"]
for i,k in enumerate(order):
    s=RF['descriptive'][k];rr=4+i;ws.cell(rr,1,k);small=('count' in k or '0-100' in k or '/KG' in k or 'per kg' in k)
    for j,key in enumerate(['count','mean','median','std','min','max']):
        c=ws.cell(rr,2+j,round(s[key],1 if small else 0));c.font=blue;c.number_format='0.0' if (small and j>0) else '#,##0'
    for c in range(1,8): ws.cell(rr,c).border=bd_
dim(ws,{'A':26,'B':8,'C':13,'D':13,'E':13,'F':12,'G':14})

# Annual Summary + native bar
ws=wb.create_sheet('Annual Summary');ws['A1']='Annual Sales vs Purchase';ws['A1'].font=title
ws.append([]);ws.append(['Financial Year','Sales (₹)','Purchase (₹)','Net (₹)']);sh(ws,3,1,4);st=4
for i,r0 in enumerate(ANN):
    rr=st+i;ws.cell(rr,1,r0['fy']);ws.cell(rr,2,round(r0['sales'])).font=blue;ws.cell(rr,3,round(r0['purchase'])).font=blue;ws.cell(rr,4,f'=B{rr}-C{rr}')
    for c in (2,3,4): ws.cell(rr,c).number_format='#,##0'
    for c in range(1,5): ws.cell(rr,c).border=bd_
last=st+len(ANN)-1
ch=BarChart();ch.type='col';ch.title='Annual Sales vs Purchase';ch.height=8;ch.width=15;ch.y_axis.title='₹'
ch.add_data(Reference(ws,min_col=2,max_col=3,min_row=3,max_row=last),titles_from_data=True);ch.set_categories(Reference(ws,min_col=1,min_row=st,max_row=last))
ch.series[0].graphicalProperties.solidFill=TEAL;ch.series[1].graphicalProperties.solidFill=AMBER
ws.add_chart(ch,'F3');dim(ws,{'A':15,'B':16,'C':16,'D':16})

# Forecast + native bar
ws=wb.create_sheet('Forecast');ws['A1']='Forecast Accuracy (6-month holdout)';ws['A1'].font=title
ws.append([]);ws.append(['Model','RMSE (₹)','MAE (₹)']);sh(ws,3,1,3);fc=RM['forecast']
for i,(m,rm,ma) in enumerate([('ARIMA',fc['sarima_rmse'],fc['sarima_mae']),('Linear (base)',fc['lin_rmse'],fc['lin_mae']),('Naive',fc['naive_rmse'],fc['naive_mae'])]):
    rr=4+i;ws.cell(rr,1,m);a=ws.cell(rr,2,round(rm));b=ws.cell(rr,3,round(ma))
    for c in (a,b): c.font=blue;c.number_format='#,##0'
    for c in range(1,4): ws.cell(rr,c).border=bd_
ch=BarChart();ch.type='col';ch.title='Forecast accuracy (lower=better)';ch.height=8;ch.width=13
ch.add_data(Reference(ws,min_col=2,max_col=3,min_row=3,max_row=6),titles_from_data=True);ch.set_categories(Reference(ws,min_col=1,min_row=4,max_row=6))
ch.series[0].graphicalProperties.solidFill=TEAL;ch.series[1].graphicalProperties.solidFill=AMBER
ws.add_chart(ch,'E3');dim(ws,{'A':16,'B':16,'C':16})

# Yield + native bar
ws=wb.create_sheet('Yield');ws['A1']='Yield per Cycle';ws['A1'].font=title
ws.append([]);ws.append(['Cycle','Yield %','Target %','Variance','Z-score','Outlier']);sh(ws,3,1,6)
for i,b in enumerate(RM['yield_batches']):
    rr=4+i;ws.cell(rr,1,f"{b['Company'][:2]}-{b['Year'][2:]}");ws.cell(rr,2,round(b['Yield_pct'],1)).font=blue;ws.cell(rr,3,22);ws.cell(rr,4,round(b['Variance'],1)).font=blue;ws.cell(rr,5,round(b['Zscore'],2)).font=blue;ws.cell(rr,6,b['Outlier'])
    for c in (2,3,4): ws.cell(rr,c).number_format='0.0'
    for c in range(1,7): ws.cell(rr,c).border=bd_
yl=4+len(RM['yield_batches'])-1
ch=BarChart();ch.type='col';ch.title='Yield % vs 22% target';ch.height=8;ch.width=15
ch.add_data(Reference(ws,min_col=2,max_col=3,min_row=3,max_row=yl),titles_from_data=True);ch.set_categories(Reference(ws,min_col=1,min_row=4,max_row=yl))
ch.series[0].graphicalProperties.solidFill=TEAL;ch.series[1].graphicalProperties.solidFill=ROSE
ws.add_chart(ch,'H3');dim(ws,{'A':10,'B':9,'C':9,'D':9,'E':9,'F':9})

# P&L Model (live formulas) + native bar of build-up
ws=wb.create_sheet('P&L Model');ws['A1']='Profit & Loss — Live Contribution Model';ws['A1'].font=title
ws['A2']='Edit blue assumption cells to test scenarios.';ws['A2'].font=ital
ws.cell(4,1,'Assumptions').font=sub;asu=RF['pnl']['assume']
ws.cell(5,1,'RCN price (₹/kg)');ws.cell(5,2,asu['rcn_price']).font=blue
ws.cell(6,1,'Outturn / yield');ws.cell(6,2,asu['outturn']).font=blue;ws.cell(6,2).number_format='0%'
ws.cell(7,1,'Processing (₹/kg kernel)');ws.cell(7,2,asu['processing_per_kg_kernel']).font=blue
ws.cell(8,1,'Kernel selling price (₹/kg)');ws.cell(8,2,asu['kernel_price']).font=blue
ws.cell(10,1,'Computed').font=sub
ws.cell(11,1,'RCN cost per kg kernel');ws.cell(11,2,'=B5/B6')
ws.cell(12,1,'Total cost per kg kernel');ws.cell(12,2,'=B11+B7')
ws.cell(13,1,'Contribution per kg kernel');ws.cell(13,2,'=B8-B12')
ws.cell(14,1,'Margin %');ws.cell(14,2,'=B13/B8');ws.cell(14,2).number_format='0.0%'
ws.cell(15,1,'Break-even price (₹/kg)');ws.cell(15,2,'=B12')
ws.cell(16,1,'Break-even outturn');ws.cell(16,2,'=B5/(B8-B7)');ws.cell(16,2).number_format='0.0%'
for r in range(5,17):
    ws.cell(r,1).border=bd_;ws.cell(r,2).border=bd_
    if ws.cell(r,2).number_format=='General':ws.cell(r,2).number_format='#,##0.0'
# build-up data for chart
ws.cell(18,1,'Profit build-up (₹/kg)').font=sub
ws.cell(19,1,'Component');ws.cell(19,2,'₹/kg');sh(ws,19,1,2)
ws.cell(20,1,'Selling price');ws.cell(20,2,'=B8')
ws.cell(21,1,'RCN cost');ws.cell(21,2,'=B11')
ws.cell(22,1,'Processing');ws.cell(22,2,'=B7')
ws.cell(23,1,'Contribution');ws.cell(23,2,'=B13')
for r in range(20,24):
    ws.cell(r,2).number_format='#,##0';ws.cell(r,1).border=bd_;ws.cell(r,2).border=bd_
ch=BarChart();ch.type='col';ch.title='Profit build-up per kg kernel';ch.height=8;ch.width=13
ch.add_data(Reference(ws,min_col=2,min_row=19,max_row=23),titles_from_data=True);ch.set_categories(Reference(ws,min_col=1,min_row=20,max_row=23))
ch.series[0].graphicalProperties.solidFill=TEAL
ws.add_chart(ch,'D4')
# grade
ws.cell(26,1,'By grade').font=sub
ws.cell(27,1,'Grade');ws.cell(27,2,'Price (₹/kg)');ws.cell(27,3,'Contribution (₹/kg)');sh(ws,27,1,3)
ws.cell(28,1,'Whole');ws.cell(28,2,RF['pnl']['whole_price']).font=blue;ws.cell(28,3,'=B28-B12')
ws.cell(29,1,'Broken');ws.cell(29,2,RF['pnl']['broken_price']).font=blue;ws.cell(29,3,'=B29-B12')
for r in (28,29):
    for c in (1,2,3): ws.cell(r,c).border=bd_
    ws.cell(r,2).number_format='#,##0';ws.cell(r,3).number_format='#,##0'
ch=BarChart();ch.type='col';ch.title='Contribution by grade (₹/kg)';ch.height=7;ch.width=11
ch.add_data(Reference(ws,min_col=3,min_row=27,max_row=29),titles_from_data=True);ch.set_categories(Reference(ws,min_col=1,min_row=28,max_row=29))
ch.series[0].graphicalProperties.solidFill=GREEN
ws.add_chart(ch,'D20');dim(ws,{'A':28,'B':14,'C':18})

# Sensitivity + colour scale (no picture)
ws=wb.create_sheet('Sensitivity');ws['A1']='Contribution Sensitivity (₹/kg kernel)';ws['A1'].font=title
ws['A2']="Rows = RCN price change · Columns = selling-price change · live formulas off 'P&L Model'.";ws['A2'].font=ital
steps=RF['pnl']['sens_steps'];lab=[f"{int(s*100):+d}%" for s in steps]
c0=ws.cell(4,1,'RCN ↓ / Price →');c0.font=hf;c0.fill=hfill;c0.alignment=center;c0.border=bd_
for j,l in enumerate(lab):
    c=ws.cell(4,2+j,l);c.font=hf;c.fill=hfill;c.alignment=center;c.border=bd_
for i,s in enumerate(steps):
    rr=5+i;c=ws.cell(rr,1,lab[i]);c.font=hf;c.fill=hfill;c.alignment=center;c.border=bd_
    for j,sp in enumerate(steps):
        f=f"='P&L Model'!$B$8*(1+{sp})-('P&L Model'!$B$5*(1+{s})/'P&L Model'!$B$6+'P&L Model'!$B$7)"
        cc2=ws.cell(rr,2+j,f);cc2.number_format='#,##0';cc2.border=bd_;cc2.alignment=center
ws.conditional_formatting.add('B5:F9',ColorScaleRule(start_type='num',start_value=-50,start_color='F8696B',mid_type='num',mid_value=60,mid_color='FFEB84',end_type='num',end_value=250,end_color='63BE7B'))
dim(ws,{'A':16,'B':10,'C':10,'D':10,'E':10,'F':10})

# Buyer Risk Model (separate sheet) + scatter + tier bar + ROC + coef bar
ws=wb.create_sheet('Buyer Risk Model');ws['A1']='Buyer Risk Model';ws['A1'].font=title
ws['A2']='Risk score 0-100 from recency, frequency, tenure and exposure. Tier sets credit limit.';ws['A2'].font=ital
hdr=['Buyer','Revenue (₹)','Orders','Avg order (₹)','Recency (d)','Tenure (d)','Months','Risk','Tier','Credit (₹)']
ws.append([]);ws.append([]);ws.append(hdr);sh(ws,4,1,10)
bds=bd.sort_values('revenue',ascending=False).reset_index(drop=True)
r0=5
for i,r in bds.iterrows():
    rr=r0+i
    ws.cell(rr,1,r.buyer);ws.cell(rr,2,round(r.revenue)).font=blue;ws.cell(rr,3,int(r.orders)).font=blue;ws.cell(rr,4,round(r.avg_order)).font=blue
    ws.cell(rr,5,int(r.recency)).font=blue;ws.cell(rr,6,int(r.tenure)).font=blue;ws.cell(rr,7,int(r.active_months)).font=blue
    ws.cell(rr,8,round(r.risk,1)).font=blue;ws.cell(rr,9,str(r.tier));ws.cell(rr,10,round(r.credit_limit)).font=blue
    for c in (2,4,10): ws.cell(rr,c).number_format='#,##0'
    ws.cell(rr,8).number_format='0.0'
    for c in range(1,11): ws.cell(rr,c).border=bd_
nb=len(bds); rl=r0+nb-1
# scatter revenue vs recency, 3 tier series -> add helper columns L..N (12..14): tier-split revenue
ws.cell(4,12,'Recency');ws.cell(4,13,'Low');ws.cell(4,14,'Medium');ws.cell(4,15,'High')
for i,r in bds.iterrows():
    rr=r0+i;ws.cell(rr,12,int(r.recency))
    ws.cell(rr,13,round(r.revenue) if r.tier=='Low' else None)
    ws.cell(rr,14,round(r.revenue) if r.tier=='Medium' else None)
    ws.cell(rr,15,round(r.revenue) if r.tier=='High' else None)
sc=ScatterChart();sc.title='Buyer risk map: revenue vs recency';sc.x_axis.title='Recency (days)';sc.y_axis.title='Revenue (₹)';sc.height=9;sc.width=16
xref=Reference(ws,min_col=12,min_row=r0,max_row=rl)
for col,colr in [(13,GREEN),(14,AMBER),(15,ROSE)]:
    yref=Reference(ws,min_col=col,min_row=4,max_row=rl)
    s=Series(yref,xref,title_from_data=True);s.marker=Marker(symbol='circle',size=6);s.graphicalProperties.line.noFill=True
    s.marker.graphicalProperties.solidFill=colr
    sc.series.append(s)
ws.add_chart(sc,'A'+str(rl+3))
# tier summary table + bar
tr=rl+3+18
ws.cell(tr,12,'Tier');ws.cell(tr,13,'Buyers');ws.cell(tr,14,'Exposure (₹)');sh(ws,tr,12,14)
for i,t in enumerate(['Low','Medium','High']):
    rr=tr+1+i;ws.cell(rr,12,t);ws.cell(rr,13,RF['buyer']['tiers'][t]).font=blue;c=ws.cell(rr,14,round(RF['buyer']['exposure'][t]));c.font=blue;c.number_format='#,##0'
    for c in range(12,15): ws.cell(rr,c).border=bd_
chb=BarChart();chb.type='col';chb.title='Buyers by tier';chb.height=7;chb.width=10
chb.add_data(Reference(ws,min_col=13,min_row=tr,max_row=tr+3),titles_from_data=True);chb.set_categories(Reference(ws,min_col=12,min_row=tr+1,max_row=tr+3))
chb.series[0].graphicalProperties.solidFill=TEAL
ws.add_chart(chb,'L'+str(tr+5))
# ROC data + scatter
roc=RF['buyer']['roc']
rocR=tr; # place ROC near col Q (17)
ws.cell(rocR,17,'FPR');ws.cell(rocR,18,'TPR');sh(ws,rocR,17,18)
for i,(f,t) in enumerate(zip(roc['fpr'],roc['tpr'])):
    ws.cell(rocR+1+i,17,round(f,4));ws.cell(rocR+1+i,18,round(t,4))
rocL=rocR+len(roc['fpr'])
scR=ScatterChart();scR.title=f"ROC — illustrative LR (AUC={RF['buyer']['lr']['auc']})";scR.x_axis.title='FPR';scR.y_axis.title='TPR';scR.height=8;scR.width=10
xr=Reference(ws,min_col=17,min_row=rocR+1,max_row=rocL);yr=Reference(ws,min_col=18,min_row=rocR,max_row=rocL)
sR=Series(yr,xr,title_from_data=True);sR.graphicalProperties.line.solidFill=TEAL;sR.graphicalProperties.line.width=22000;sR.marker=Marker(symbol='none');scR.series.append(sR)
ws.add_chart(scR,'Q'+str(rocL+2))
# coefficients bar
co=RF['buyer']['lr']['coef'];coR=rocL+2
ws.cell(coR,17,'Feature');ws.cell(coR,18,'Coef');sh(ws,coR,17,18)
for i,(k,v) in enumerate(co.items()):
    ws.cell(coR+1+i,17,k);ws.cell(coR+1+i,18,round(v,3))
coL=coR+len(co)
chc=BarChart();chc.type='bar';chc.title='Delay-risk drivers (illustrative)';chc.height=7;chc.width=10
chc.add_data(Reference(ws,min_col=18,min_row=coR,max_row=coL),titles_from_data=True);chc.set_categories(Reference(ws,min_col=17,min_row=coR+1,max_row=coL))
chc.series[0].graphicalProperties.solidFill=AMBER
ws.add_chart(chc,'L'+str(coR+1))
ws.cell(2,12,f"Illustrative logistic model: AUC={RF['buyer']['lr']['auc']}, accuracy={RF['buyer']['lr']['acc']} (simulated delays).").font=ital
dim(ws,{'A':28,'B':14,'C':7,'D':12,'E':10,'F':9,'G':8,'H':7,'I':8,'J':13})

# Buyer Tiers small sheet
ws=wb.create_sheet('Buyer Tiers');ws['A1']='Buyer Tiers & Credit Exposure';ws['A1'].font=title
ws.append([]);ws.append(['Tier','Buyers','Credit exposure (₹)']);sh(ws,3,1,3)
for i,t in enumerate(['Low','Medium','High']):
    rr=4+i;ws.cell(rr,1,t);ws.cell(rr,2,RF['buyer']['tiers'][t]).font=blue;c=ws.cell(rr,3,round(RF['buyer']['exposure'][t]));c.font=blue;c.number_format='#,##0'
    for c in range(1,4): ws.cell(rr,c).border=bd_
ch=BarChart();ch.type='col';ch.title='Credit exposure by tier';ch.height=8;ch.width=12
ch.add_data(Reference(ws,min_col=3,min_row=3,max_row=6),titles_from_data=True);ch.set_categories(Reference(ws,min_col=1,min_row=4,max_row=6));ch.series[0].graphicalProperties.solidFill=TEAL
ws.add_chart(ch,'E3');dim(ws,{'A':10,'B':9,'C':18})

# Anomalies + scatter (normal vs flagged) + reconciliation line
ws=wb.create_sheet('Anomalies');ws['A1']='Anomaly Detection';ws['A1'].font=title
an=RF['anomaly']
ws.append([]);ws.cell(3,1,f"Flagged {an['n_total']} of {RF['totals']['n_sales']+RF['totals']['n_purchase']} ({an['pct']}%): {an['n_zscore']} Z-score, {an['n_iso']} Isolation Forest. {an['recon_months_flagged']} months flagged by reconciliation.").font=ital
ws.append([]);ws.append(['Date','Company','Flow','Goods','Party','Amount (₹)','Z','Flagged by']);sh(ws,5,1,8)
for i,a in enumerate(an['top']):
    rr=6+i;ws.cell(rr,1,a['Date']);ws.cell(rr,2,a['Company']);ws.cell(rr,3,a['Flow']);ws.cell(rr,4,a['Goods']);ws.cell(rr,5,a['Party']);c=ws.cell(rr,6,round(a['Amount']));c.font=blue;c.number_format='#,##0';ws.cell(rr,7,a['Z']);ws.cell(rr,8,a['Reason'])
    for c in range(1,9): ws.cell(rr,c).border=bd_
# scatter data: index, normal amount, anomaly amount  (cols J..L =10..12)
ws.cell(5,10,'Idx');ws.cell(5,11,'Normal');ws.cell(5,12,'Anomaly');sh(ws,5,10,12)
Ar=A.reset_index(drop=True)
for i,r in Ar.iterrows():
    rr=6+i;ws.cell(rr,10,i+1)
    if r.anomaly: ws.cell(rr,12,round(r.Amount))
    else: ws.cell(rr,11,round(r.Amount))
al=6+len(Ar)-1
scA=ScatterChart();scA.title='Transactions — anomalies flagged';scA.x_axis.title='transaction #';scA.y_axis.title='Amount (₹)';scA.height=9;scA.width=17
xr=Reference(ws,min_col=10,min_row=6,max_row=al)
for col,colr,nm in [(11,'CBD5E1','Normal'),(12,ROSE,'Anomaly')]:
    yr=Reference(ws,min_col=col,min_row=5,max_row=al);s=Series(yr,xr,title_from_data=True);s.marker=Marker(symbol='circle',size=4 if col==12 else 3);s.graphicalProperties.line.noFill=True;s.marker.graphicalProperties.solidFill=colr;scA.series.append(s)
ws.add_chart(scA,'A20')
# reconciliation line (cols N..O =14..15)
ws.cell(5,14,'Month');ws.cell(5,15,'Kernel/ RCN ratio');sh(ws,5,14,15)
rc=recon.reset_index()
for i,r in rc.iterrows():
    rr=6+i;ws.cell(rr,14,pd.Timestamp(r['Date']).strftime('%Y-%m'));ws.cell(rr,15,round(float(r['ratio']),3))
rl2=6+len(rc)-1
lc=LineChart();lc.title='Stock reconciliation — monthly ratio';lc.height=8;lc.width=16
lc.add_data(Reference(ws,min_col=15,min_row=5,max_row=rl2),titles_from_data=True);lc.set_categories(Reference(ws,min_col=14,min_row=6,max_row=rl2));lc.series[0].graphicalProperties.line.solidFill=TEAL
ws.add_chart(lc,'N20')
dim(ws,{'A':12,'B':9,'C':8,'D':7,'E':24,'F':13,'G':7,'H':15})

wb.save('Ankita_Adrita_Final_Analysis.xlsx')
print('saved native-chart workbook:',wb.sheetnames)
